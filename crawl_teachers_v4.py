#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
浙大导师信息爬取脚本 v4 - 简化版
直接从HTML结构提取研究方向
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 导师列表页面URL
LIST_URL = "https://pi.zju.edu.cn/2026/0317/c90502a3141478/page.htm"

# 输出路径
OUTPUT_DIR = "/home/admin/.openclaw/workspace"
EXCEL_PATH = os.path.join(OUTPUT_DIR, "导师信息_v4.xlsx")
CV_DIR = os.path.join(OUTPUT_DIR, "导师简历_v4")
REPORT_PATH = os.path.join(OUTPUT_DIR, "导师分析报告_v4.md")

os.makedirs(CV_DIR, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

# ============================================================
# 分类体系
# ============================================================

HARD_DISCIPLINES = {
    "集成电路设计": {
        "模拟IC设计": ["模拟集成电路", "模拟ic", "模拟电路", "adc", "dac", "运放", "放大器", "电源管理芯片", "pmic"],
        "射频IC设计": ["射频集成电路", "射频ic", "rfic", "射频电路", "毫米波电路", "收发机", "锁相环", "频率合成"],
        "数字IC设计": ["数字集成电路", "数字ic", "vlsi", "soc", "数字电路", "逻辑设计", "前端设计", "后端设计"],
        "混合信号IC": ["混合信号", "数模混合", "mixed-signal"],
    },
    "半导体器件": {
        "功率半导体": ["功率器件", "功率半导体", "gan", "氮化镓", "sic", "碳化硅", "igbt", "功率模块"],
        "微电子工艺": ["微电子工艺", "半导体工艺", "制造工艺", "光刻", "刻蚀", "薄膜", "封装", "测试"],
        "新型器件": ["新型半导体器件", "忆阻器", "铁电器件", "相变存储", "阻变存储"],
    },
    "MEMS与微系统": {
        "MEMS传感器": ["mems", "微机电", "微纳", "微系统", "惯性传感器", "压力传感器", "mems器件"],
        "生物MEMS": ["生物微机电", "微流控", "生物芯片", "lab-on-chip"],
    },
    "光电子器件": {
        "光电子器件": ["光电子", "光电器件", "光电探测器", "激光器", "led", "光伏", "光通信器件"],
        "集成光子学": ["集成光子", "硅光子", "光子芯片", "光波导", "光子器件"],
        "显示技术": ["显示", "oled", "lcd", "量子点", "microled", "显示面板"],
    },
    "电磁与射频硬件": {
        "天线设计": ["天线", "阵列天线", "相控阵", "mimo天线", "超表面天线"],
        "射频电路与器件": ["射频电路", "微波电路", "滤波器", "功分器", "耦合器"],
        "电磁材料": ["人工电磁材料", "超材料", "电磁屏蔽", "吸波材料"],
    },
}

SOFT_DISCIPLINES = {
    "人工智能": {
        "深度学习理论": ["深度学习理论", "神经网络理论", "优化算法", "学习理论", "表示学习"],
        "计算机视觉": ["计算机视觉", "图像处理", "目标检测", "图像分割", "视觉感知", "三维视觉", "多光谱成像", "图像识别"],
        "自然语言处理": ["自然语言处理", "nlp", "大语言模型", "llm", "文本挖掘", "机器翻译", "对话系统"],
        "语音与音频处理": ["语音识别", "语音合成", "声学信号", "语音处理", "音频处理"],
        "机器学习应用": ["机器学习", "深度学习应用", "ai应用", "智能算法", "数据挖掘"],
        "强化学习": ["强化学习", "决策智能", "智能控制", "博弈论"],
        "生成式AI": ["生成模型", "diffusion", "gan", "生成式人工智能", "aigc"],
        "AI系统与加速": ["ai加速", "深度学习加速", "推理加速", "边缘ai"],
        "多模态学习": ["多模态", "视觉语言", "跨模态学习"],
    },
    "信号处理": {
        "数字信号处理": ["数字信号处理", "dsp", "信号处理", "时频分析", "自适应滤波"],
        "图像与视频处理": ["图像处理", "视频处理", "视频编码", "图像压缩", "视频分析"],
        "通信信号处理": ["通信信号处理", "调制解调", "信道编码", "信道估计", "ofdm"],
    },
    "通信网络": {
        "无线通信算法": ["无线通信", "5g", "6g", "移动通信", "蜂窝网络", "通信算法"],
        "网络技术": ["网络", "物联网", "iot", "边缘计算", "云计算", "sdn"],
        "通信理论": ["信息论", "编码理论", "网络信息论"],
    },
    "智能系统应用": {
        "机器人与导航": ["机器人", "导航定位", "slam", "无人系统", "自主导航"],
        "智能感知": ["智能感知", "环境感知", "传感器融合", "目标跟踪"],
        "智能控制": ["智能控制", "自适应控制", "运动控制"],
    },
}

INTERDISCIPLINARY = {
    "量子技术": {
        "量子计算": ["量子计算", "量子算法", "量子电路"],
        "量子通信": ["量子通信", "量子密钥"],
        "量子器件": ["量子光源", "量子芯片", "量子点"],
    },
    "类脑计算": {
        "类脑器件": ["类脑计算", "神经形态器件", "脉冲神经网络"],
        "脑机接口": ["脑机接口", "bci", "神经接口"],
    },
    "EDA与设计工具": {
        "EDA工具": ["eda", "电子设计自动化", "版图设计", "物理验证"],
        "电路仿真": ["电路仿真", "电磁仿真", "多物理仿真"],
    },
}


def build_keyword_mapping():
    mapping = {}
    for cat1, subcats in HARD_DISCIPLINES.items():
        for cat2, keywords in subcats.items():
            for kw in keywords:
                mapping[kw.lower()] = ("硬学科", cat1, cat2)
    for cat1, subcats in SOFT_DISCIPLINES.items():
        for cat2, keywords in subcats.items():
            for kw in keywords:
                mapping[kw.lower()] = ("软学科", cat1, cat2)
    for cat1, subcats in INTERDISCIPLINARY.items():
        for cat2, keywords in subcats.items():
            for kw in keywords:
                mapping[kw.lower()] = ("交叉学科", cat1, cat2)
    return mapping

KEYWORD_MAPPING = build_keyword_mapping()


def classify_research(research_text):
    if not research_text or research_text == "未获取":
        return "未分类", "其他", "其他", research_text
    
    text_lower = research_text.lower()
    scores = {}
    
    for keyword, (type1, cat1, cat2) in KEYWORD_MAPPING.items():
        if keyword in text_lower:
            key = (type1, cat1, cat2)
            if key not in scores:
                scores[key] = 0
            scores[key] += len(keyword)
    
    if scores:
        best = max(scores.items(), key=lambda x: x[1])
        type1, cat1, cat2 = best[0]
        return type1, cat1, cat2, research_text
    
    hard_keywords = ["ic", "芯片", "集成电路", "vlsi", "半导体", "器件", "工艺", "mems", 
                     "天线", "射频电路", "微波电路", "光电器件", "激光器", "封装"]
    soft_keywords = ["机器学习", "深度学习", "人工智能", "ai", "视觉", "图像", "算法",
                     "信号处理", "通信", "网络", "控制", "机器人", "软件"]
    
    hard_score = sum(1 for kw in hard_keywords if kw in text_lower)
    soft_score = sum(1 for kw in soft_keywords if kw in text_lower)
    
    if hard_score > soft_score:
        return "硬学科", "集成电路与硬件", "其他硬件", research_text
    elif soft_score > hard_score:
        return "软学科", "信号与智能处理", "其他软件", research_text
    
    return "未分类", "其他", "其他", research_text


def get_teacher_list():
    print("正在获取导师列表...")
    try:
        response = requests.get(LIST_URL, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        
        teachers = []
        table = soup.find('table')
        if table:
            rows = table.find_all('tr')[1:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 3:
                    seq = cols[0].get_text(strip=True)
                    name_link = cols[1].find('a')
                    if name_link:
                        name = name_link.get_text(strip=True)
                        url = name_link.get('href', '')
                        if url and not url.startswith('http'):
                            url = 'https://person.zju.edu.cn' + url if url.startswith('/') else 'https://person.zju.edu.cn/' + url
                    else:
                        name = cols[1].get_text(strip=True)
                        url = ''
                    college = cols[2].get_text(strip=True)
                    
                    teachers.append({
                        'seq': seq,
                        'name': name,
                        'college': college,
                        'url': url
                    })
        
        print(f"获取到 {len(teachers)} 位导师")
        return teachers
    except Exception as e:
        print(f"获取导师列表失败: {e}")
        return []


def get_teacher_detail(url, name):
    """获取单个导师详细信息 - 直接从HTML提取"""
    if not url:
        return {
            'name': name,
            'title': '未获取',
            'research': '未获取',
            'bio': '未获取',
            'email': '未获取'
        }
    
    try:
        print(f"  正在获取: {name}")
        response = requests.get(url, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取职称 - 从页面文本中提取
        title = "未获取"
        page_text = soup.get_text()
        title_match = re.search(r'(教授|副教授|讲师|研究员|副研究员|高级工程师|特聘研究员|青年学者|百人计划|长聘|特聘)', page_text)
        if title_match:
            title = title_match.group(1)
        
        # 提取研究方向 - 从 second_research 类中提取
        research = "未获取"
        research_list = []
        
        # 方法1: 从 second_research 类提取
        research_ul = soup.find('ul', class_='second_research')
        if research_ul:
            li_items = research_ul.find_all('li')
            for li in li_items:
                text = li.get_text(strip=True)
                # 移除前面的点号
                text = text.strip('·• \xa0')
                if text and len(text) > 1:
                    research_list.append(text)
        
        # 方法2: 如果方法1没有找到，尝试其他结构
        if not research_list:
            # 尝试查找包含"研究方向"的区域
            for elem in soup.find_all(['ul', 'div']):
                text = elem.get_text()
                if '研究方向' in text and '·' in text:
                    # 提取列表项
                    items = re.findall(r'·\s*([^·\n]+)', text)
                    for item in items:
                        item = item.strip()
                        if item and len(item) > 1 and len(item) < 50:
                            research_list.append(item)
                    if research_list:
                        break
        
        if research_list:
            research = "；".join(research_list)
        
        # 提取邮箱
        email = "未获取"
        mailto = soup.find('a', href=re.compile(r'^mailto:'))
        if mailto:
            email = mailto.get('href').replace('mailto:', '')
        else:
            email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', page_text)
            if email_match:
                email = email_match.group(0)
        
        # 提取个人简介 - 从个人简介区域提取
        bio = "未获取"
        # 查找个人简介区域
        bio_div = soup.find('div', class_='content') or soup.find('div', class_='bio')
        if bio_div:
            bio = bio_div.get_text(strip=True)[:500]
        
        return {
            'name': name,
            'title': title,
            'research': research,
            'bio': bio,
            'email': email
        }
        
    except Exception as e:
        print(f"    获取 {name} 详情失败: {e}")
        return {
            'name': name,
            'title': '获取失败',
            'research': '获取失败',
            'bio': '获取失败',
            'email': '获取失败'
        }


def save_cv(teacher):
    filename = f"{teacher['seq']}_{teacher['name']}.md"
    filepath = os.path.join(CV_DIR, filename)
    
    content = f"""# {teacher['name']}

## 基本信息
- **姓名**: {teacher['name']}
- **学院**: {teacher['college']}
- **职称**: {teacher['title']}
- **邮箱**: {teacher['email']}
- **个人页面**: {teacher['url']}

## 研究方向
{teacher['research']}

## 个人简介
{teacher['bio']}

## 分类信息
- **学科类型**: {teacher['type1']}
- **大类**: {teacher['cat1']}
- **中类**: {teacher['cat2']}

---
*数据来源: 浙江大学导师信息系统*
*爬取时间: {time.strftime('%Y-%m-%d %H:%M:%S')}*
"""
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)


def create_excel(teachers_data):
    print("\n正在创建Excel文件...")
    
    all_data = []
    for t in teachers_data:
        type1, cat1, cat2, research = classify_research(t['research'])
        t['type1'] = type1
        t['cat1'] = cat1
        t['cat2'] = cat2
        
        all_data.append({
            '序号': t['seq'],
            '姓名': t['name'],
            '学院': t['college'],
            '职称': t['title'],
            '学科类型': type1,
            '研究方向大类': cat1,
            '研究方向中类': cat2,
            '具体研究方向': t['research'],
            '邮箱': t['email'],
            '个人页面': t['url'],
        })
    
    df_all = pd.DataFrame(all_data)
    
    wb = Workbook()
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hard_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    soft_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 全部导师
    ws_all = wb.active
    ws_all.title = "全部导师"
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
            elif r_idx > 1 and c_idx == 5:
                if value == "硬学科":
                    cell.fill = hard_fill
                elif value == "软学科":
                    cell.fill = soft_fill
    
    column_widths = {'A': 6, 'B': 10, 'C': 18, 'D': 12, 'E': 10, 'F': 18, 'G': 15, 'H': 50, 'I': 25, 'J': 40}
    for col, width in column_widths.items():
        ws_all.column_dimensions[col].width = width
    
    # 软学科
    ws_soft = wb.create_sheet(title="软学科导师")
    df_soft = df_all[df_all['学科类型'] == '软学科']
    for r_idx, row in enumerate(dataframe_to_rows(df_soft, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_soft.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
    
    # 硬学科
    ws_hard = wb.create_sheet(title="硬学科导师")
    df_hard = df_all[df_all['学科类型'] == '硬学科']
    for r_idx, row in enumerate(dataframe_to_rows(df_hard, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_hard.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
    
    # 统计
    ws_summary = wb.create_sheet(title="统计汇总")
    ws_summary.cell(row=1, column=1, value="学科类型统计").font = Font(bold=True, size=14)
    
    type_counts = df_all['学科类型'].value_counts()
    for i, (t, count) in enumerate(type_counts.items(), 3):
        ws_summary.cell(row=i, column=1, value=t)
        ws_summary.cell(row=i, column=2, value=count)
        ws_summary.cell(row=i, column=3, value=f"{count/len(df_all)*100:.1f}%")
    
    # AI细分
    df_ai = df_all[df_all['研究方向大类'] == '人工智能']
    if len(df_ai) > 0:
        ws_summary.cell(row=10, column=1, value="人工智能细分").font = Font(bold=True, size=14)
        ai_cat2 = df_ai['研究方向中类'].value_counts()
        for i, (cat, count) in enumerate(ai_cat2.items(), 11):
            ws_summary.cell(row=i, column=1, value=cat)
            ws_summary.cell(row=i, column=2, value=count)
    
    wb.save(EXCEL_PATH)
    print(f"Excel文件已保存: {EXCEL_PATH}")
    
    return df_all


def generate_report(teachers_data, df_all):
    print("\n正在生成分析报告...")
    
    total = len(teachers_data)
    type_counts = df_all['学科类型'].value_counts()
    df_soft = df_all[df_all['学科类型'] == '软学科']
    df_hard = df_all[df_all['学科类型'] == '硬学科']
    df_ai = df_all[df_all['研究方向大类'] == '人工智能']
    
    soft_cat1_counts = df_soft['研究方向大类'].value_counts()
    hard_cat1_counts = df_hard['研究方向大类'].value_counts()
    ai_cat2_counts = df_ai['研究方向中类'].value_counts()
    
    report = f"""# 浙江大学电子信息技术方向导师分析报告

> 生成时间: {time.strftime('%Y年%m月%d日')}
> 数据来源: https://pi.zju.edu.cn/2026/0317/c90502a3141478/page.htm
> 导师总数: {total}位

---

## 一、学科类型总览（软学科 vs 硬学科）

### 1.1 什么是软学科 vs 硬学科？

**软学科（Software-oriented）** 🟢：
- 侧重算法、软件、数据处理、应用开发
- 通常可以在普通计算机上开展研究
- 不依赖昂贵的硬件设备和实验室设施
- 适合：编程能力强、数学基础好、希望快速迭代的学生

**硬学科（Hardware-oriented）** 🔴：
- 需要硬件设施、实验室设备、芯片流片等
- 往往需要专业实验室、测试仪器
- 研究周期较长，成果转化需要硬件验证
- 适合：动手能力强、对硬件感兴趣、有实验室资源的学生

### 1.2 学科类型分布

| 学科类型 | 导师数量 | 占比 | 入门难度 | 推荐指数 |
|---------|---------|------|---------|---------|
"""
    
    for t, count in type_counts.items():
        if t == "软学科":
            desc = "⭐ 较低 | 💪💪💪 强烈推荐"
        elif t == "硬学科":
            desc = "⭐⭐⭐ 较高 | 💪💪 推荐"
        elif t == "交叉学科":
            desc = "⭐⭐ 中等 | 💪💪 推荐"
        else:
            desc = "待分析 | 待分析"
        report += f"| {t} | {count} | {count/total*100:.1f}% | {desc} |\n"
    
    report += f"""
---

## 二、软学科详细分析（重点推荐）

### 2.1 为什么优先推荐软学科？

✅ **入门门槛相对较低**：通常只需要一台电脑和编程能力
✅ **研究资源易获取**：开源数据集、框架、工具丰富
✅ **成果转化快**：论文、代码、应用可以快速产出
✅ **就业面广**：互联网、金融、科技公司都需要
✅ **适合远程研究**：不需要实验室也能做出成果
✅ **适合你的情况**：无法接触集成电路等硬件设备

### 2.2 软学科大类分布

| 研究方向大类 | 导师数量 | 占软学科比例 | 就业前景 |
|-------------|---------|-------------|---------|
"""
    
    for cat, count in soft_cat1_counts.sort_values(ascending=False).items():
        report += f"| {cat} | {count} | {count/len(df_soft)*100:.1f}% | 💰💰💰 |\n"
    
    # AI详细分析
    report += f"""
---

## 三、人工智能方向细化分析（重点）

### 3.1 AI细分领域分布

| 细分领域 | 导师数量 | 研究内容 | 就业前景 | 推荐度 |
|---------|---------|---------|---------|-------|
"""
    
    ai_subfields = {
        "计算机视觉": ("图像处理、目标检测、视频分析、视觉感知", "💰💰💰", "⭐⭐⭐⭐⭐"),
        "深度学习理论": ("神经网络理论、优化算法、表示学习", "💰💰💰", "⭐⭐⭐⭐"),
        "机器学习应用": ("数据挖掘、智能算法、AI应用开发", "💰💰💰", "⭐⭐⭐⭐⭐"),
        "自然语言处理": ("大语言模型、文本分析、对话系统", "💰💰💰💰", "⭐⭐⭐⭐⭐"),
        "生成式AI": ("Diffusion模型、GAN、AIGC应用", "💰💰💰💰", "⭐⭐⭐⭐⭐"),
        "AI系统与加速": ("深度学习加速、推理优化、边缘AI", "💰💰💰", "⭐⭐⭐⭐"),
        "语音与音频处理": ("语音识别、语音合成、声学信号", "💰💰💰", "⭐⭐⭐"),
        "强化学习": ("决策智能、智能控制、博弈论", "💰💰💰", "⭐⭐⭐"),
        "多模态学习": ("视觉语言、跨模态理解", "💰💰💰💰", "⭐⭐⭐⭐"),
    }
    
    for cat, count in ai_cat2_counts.sort_values(ascending=False).items():
        info = ai_subfields.get(cat, ("相关研究", "💰💰💰", "⭐⭐⭐"))
        report += f"| {cat} | {count} | {info[0]} | {info[1]} | {info[2]} |\n"
    
    # AI导师详细列表
    if len(df_ai) > 0:
        report += """
### 3.2 AI各细分领域导师名单

"""
        
        for cat2 in ai_cat2_counts.sort_values(ascending=False).index:
            df_sub = df_ai[df_ai['研究方向中类'] == cat2]
            report += f"#### {cat2} ({len(df_sub)}人)\n\n"
            report += "| 序号 | 姓名 | 职称 | 学院 | 具体研究方向 |\n|------|------|------|------|-------------|\n"
            for _, row in df_sub.iterrows():
                research = str(row['具体研究方向'])[:40] if row['具体研究方向'] else "-"
                report += f"| {row['序号']} | {row['姓名']} | {row['职称']} | {row['学院'][:8]} | {research} |\n"
            report += "\n"
    
    # 其他软学科
    report += """
---

## 四、软学科其他方向

### 4.1 信号处理方向

"""
    
    df_signal = df_soft[df_soft['研究方向大类'] == '信号处理']
    if len(df_signal) > 0:
        signal_cat2 = df_signal['研究方向中类'].value_counts()
        for cat, count in signal_cat2.items():
            report += f"**{cat}** ({count}人)\n\n"
            for _, row in df_signal[df_signal['研究方向中类'] == cat].iterrows():
                research = str(row['具体研究方向'])[:40] if row['具体研究方向'] else "-"
                report += f"- {row['姓名']} ({row['职称']}): {research}\n"
            report += "\n"
    
    report += """
### 4.2 通信网络方向

"""
    
    df_comm = df_soft[df_soft['研究方向大类'] == '通信网络']
    if len(df_comm) > 0:
        comm_cat2 = df_comm['研究方向中类'].value_counts()
        for cat, count in comm_cat2.items():
            report += f"**{cat}** ({count}人)\n\n"
            for _, row in df_comm[df_comm['研究方向中类'] == cat].iterrows():
                research = str(row['具体研究方向'])[:40] if row['具体研究方向'] else "-"
                report += f"- {row['姓名']} ({row['职称']}): {research}\n"
            report += "\n"
    
    # 硬学科概览
    report += f"""
---

## 五、硬学科概览

### 5.1 硬学科特点

⚠️ **需要实验室资源**：芯片设计需要EDA工具和流片机会
⚠️ **研究周期长**：从设计到验证可能需要数月甚至数年
⚠️ **专业性强**：需要深入的硬件知识
⚠️ **高门槛高回报**：入行难，但人才稀缺，薪资高

### 5.2 硬学科大类分布

| 研究方向大类 | 导师数量 | 入门难度 | 薪资前景 |
|-------------|---------|---------|---------|
"""
    
    for cat, count in hard_cat1_counts.sort_values(ascending=False).items():
        report += f"| {cat} | {count} | ⭐⭐⭐ | 💰💰💰💰 |\n"
    
    # 硬学科导师名单
    if len(df_hard) > 0:
        report += """
### 5.3 硬学科导师名单

"""
        
        for cat1 in hard_cat1_counts.sort_values(ascending=False).index:
            df_sub = df_hard[df_hard['研究方向大类'] == cat1]
            report += f"#### {cat1} ({len(df_sub)}人)\n\n"
            for _, row in df_sub.iterrows():
                research = str(row['具体研究方向'])[:40] if row['具体研究方向'] else "-"
                report += f"- {row['姓名']} ({row['职称']}): {row['研究方向中类']} - {research}\n"
            report += "\n"
    
    # 推荐
    report += f"""
---

## 六、针对你的情况的选导师建议

考虑到你提到**可能无法接触集成电路等硬学科**，以下是针对你的**软学科优先推荐**：

### 6.1 第一梯队（强烈推荐）

**计算机视觉**：
- ✅ 入门相对容易，教程和资源丰富
- ✅ 就业需求大，互联网、汽车、安防都需要
- ✅ 可以远程研究，不需要特殊设备
- 推荐导师：沈会良、杜歆、马蔚、龚小谨等

**自然语言处理/大语言模型**：
- ✅ 当前最热门方向
- ✅ 开源模型丰富，可以快速上手
- ✅ 就业前景极好
- 注意：竞争激烈，需要强编程能力

**生成式AI（Diffusion/GAN/AIGC）**：
- ✅ 新兴热门方向
- ✅ 应用场景广泛
- ✅ 创业机会多

### 6.2 第二梯队（稳定选择）

**机器学习应用**：
- ✅ 就业面最广
- ✅ 各行业都需要
- ✅ 相对不那么卷

**信号处理**：
- ✅ 传统优势方向
- ✅ 通信、音频、图像处理都有应用
- ✅ 相对稳定

### 6.3 选导师实用建议

1. **先看研究方向是否感兴趣**
2. **查看导师近3年的论文和项目**
3. **了解实验室的毕业要求和氛围**
4. **提前联系，了解是否招人**
5. **如果能联系到师兄师姐，多了解实验室情况**

---

## 七、数据统计摘要

| 指标 | 数值 |
|------|------|
| 导师总数 | {total} |
| 软学科导师 | {len(df_soft)} ({len(df_soft)/total*100:.1f}%) |
| 硬学科导师 | {len(df_hard)} ({len(df_hard)/total*100:.1f}%) |
| AI相关导师 | {len(df_ai)} ({len(df_ai)/total*100:.1f}%) |
| 教授/研究员 | {len(df_all[df_all['职称'].str.contains('教授|研究员', na=False)])} |

---

## 附录：完整导师名单

### A. 软学科导师完整名单

| 序号 | 姓名 | 学院 | 职称 | 大类 | 中类 | 研究方向 |
|------|------|------|------|------|------|----------|
"""
    
    for _, row in df_soft.sort_values(['研究方向大类', '研究方向中类']).iterrows():
        research = str(row['具体研究方向'])[:30] if row['具体研究方向'] else "-"
        report += f"| {row['序号']} | {row['姓名']} | {row['学院'][:8]} | {row['职称']} | {row['研究方向大类'][:6]} | {row['研究方向中类'][:6]} | {research} |\n"
    
    report += """
### B. 硬学科导师完整名单

| 序号 | 姓名 | 学院 | 职称 | 大类 | 中类 |
|------|------|------|------|------|------|
"""
    
    for _, row in df_hard.sort_values(['研究方向大类', '研究方向中类']).iterrows():
        report += f"| {row['序号']} | {row['姓名']} | {row['学院'][:8]} | {row['职称']} | {row['研究方向大类'][:6]} | {row['研究方向中类'][:6]} |\n"
    
    report += f"""

---

*本报告由AI自动生成*
*生成时间: {time.strftime('%Y年%m月%d日 %H:%M:%S')}*
"""
    
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"分析报告已保存: {REPORT_PATH}")


def main():
    print("=" * 60)
    print("浙江大学电子信息技术方向导师信息爬取 v4")
    print("简化版 - 直接从HTML结构提取研究方向")
    print("=" * 60)
    
    # 获取导师列表
    teachers = get_teacher_list()
    if not teachers:
        print("获取导师列表失败，退出")
        return
    
    # 获取每位导师详细信息
    print(f"\n开始爬取 {len(teachers)} 位导师的详细信息...")
    teachers_data = []
    
    for i, t in enumerate(teachers):
        print(f"[{i+1}/{len(teachers)}] 处理: {t['name']}")
        
        detail = get_teacher_detail(t['url'], t['name'])
        
        teacher_data = {
            'seq': t['seq'],
            'name': t['name'],
            'college': t['college'],
            'url': t['url'],
            **detail
        }
        teachers_data.append(teacher_data)
        
        time.sleep(1.5)
    
    # 创建Excel
    df_all = create_excel(teachers_data)
    
    # 保存简历
    print("\n正在保存导师简历...")
    for t in teachers_data:
        save_cv(t)
    
    # 生成报告
    generate_report(teachers_data, df_all)
    
    print("\n" + "=" * 60)
    print("爬取完成！")
    print(f"Excel文件: {EXCEL_PATH}")
    print(f"简历目录: {CV_DIR}")
    print(f"分析报告: {REPORT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()