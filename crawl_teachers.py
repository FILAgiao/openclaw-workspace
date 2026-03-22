#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
浙大导师信息爬取脚本
从导师列表页面获取详细信息并生成Excel和分析报告
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 导师列表页面URL
LIST_URL = "https://pi.zju.edu.cn/2026/0317/c90502a3141478/page.htm"

# 输出路径
OUTPUT_DIR = "/home/admin/.openclaw/workspace"
EXCEL_PATH = os.path.join(OUTPUT_DIR, "导师信息.xlsx")
CV_DIR = os.path.join(OUTPUT_DIR, "导师简历")
REPORT_PATH = os.path.join(OUTPUT_DIR, "导师分析报告.md")

# 请求头
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Connection': 'keep-alive',
}

# 创建简历目录
os.makedirs(CV_DIR, exist_ok=True)

# 研究方向分类体系
CATEGORY_SYSTEM = {
    "集成电路与芯片": {
        "集成电路设计": ["模拟IC", "数字IC", "射频IC", "功率IC", "混合信号", "集成电路", "IC设计", "芯片设计"],
        "半导体器件": ["半导体", "器件物理", "晶体管", "功率器件", "宽禁带", "氮化镓", "碳化硅", "SiC", "GaN"],
        "MEMS传感器": ["MEMS", "微机电", "传感器", "微纳", "微系统"],
        "EDA工具": ["EDA", "电子设计自动化", "版图", "验证", "综合"],
        "封装测试": ["封装", "测试", "可靠性", "工艺"]
    },
    "通信与网络": {
        "无线通信": ["5G", "6G", "无线", "移动通信", "蜂窝", "OFDM", "MIMO", "毫米波"],
        "光通信": ["光通信", "光纤", "光网络", "波分复用", "光传输"],
        "网络技术": ["网络", "物联网", "IoT", "边缘计算", "云计算", "SDN", "网络安全"],
        "通信信号处理": ["通信信号", "调制解调", "信道编码", "信道估计"]
    },
    "电磁与微波": {
        "射频技术": ["射频", "RF", "微波", "射频电路", "天线", "波束成形"],
        "电磁场理论": ["电磁场", "电磁波", "电磁兼容", "电磁散射", "计算电磁"],
        "微波器件": ["微波器件", "滤波器", "放大器", "混频器", "振荡器"],
        "天线技术": ["天线", "阵列天线", "相控阵", "MIMO天线", "超表面"]
    },
    "信号处理与人工智能": {
        "机器学习": ["机器学习", "深度学习", "神经网络", "人工智能", "AI", "强化学习", "学习算法"],
        "计算机视觉": ["计算机视觉", "图像处理", "图像识别", "目标检测", "图像分割", "视觉"],
        "语音信号处理": ["语音", "语音识别", "语音合成", "声学", "声音"],
        "信号处理": ["信号处理", "数字信号", "DSP", "时频分析", "变换", "滤波"],
        "数据科学": ["大数据", "数据挖掘", "数据科学", "统计分析"]
    },
    "光电子与显示": {
        "光电子器件": ["光电子", "光电", "激光器", "探测器", "光电器件", "光电探测"],
        "显示技术": ["显示", "OLED", "LCD", "量子点", "显示面板", "显示屏"],
        "光子学": ["光子", "光子学", "纳米光子", "集成光子", "硅光子"],
        "照明技术": ["LED", "照明", "固态照明"]
    },
    "电子系统与嵌入式": {
        "嵌入式系统": ["嵌入式", "单片机", "MCU", "固件", "实时系统", "嵌入式软件"],
        "电路系统": ["电路系统", "电路设计", "PCB", "硬件", "电路板"],
        "电子测量": ["测量", "测试", "仪器", "检测", "传感"],
        "智能系统": ["智能系统", "智能控制", "机器人", "自动化", "控制"]
    },
    "其他": {
        "交叉学科": ["交叉", "融合", "多学科"],
        "其他": []
    }
}

def build_keyword_mapping():
    """建立关键词到分类的映射"""
    mapping = {}
    for cat1, subcats in CATEGORY_SYSTEM.items():
        for cat2, keywords in subcats.items():
            for kw in keywords:
                mapping[kw.lower()] = (cat1, cat2)
    return mapping

KEYWORD_MAPPING = build_keyword_mapping()

def classify_research(research_text):
    """根据研究方向文本进行分类"""
    if not research_text or research_text == "未获取":
        return "其他", "其他", research_text
    
    text_lower = research_text.lower()
    scores = {}  # (cat1, cat2) -> score
    
    for keyword, (cat1, cat2) in KEYWORD_MAPPING.items():
        if keyword in text_lower:
            key = (cat1, cat2)
            scores[key] = scores.get(key, 0) + len(keyword)  # 长关键词权重更高
    
    if scores:
        # 找出得分最高的分类
        best = max(scores.items(), key=lambda x: x[1])
        cat1, cat2 = best[0]
        return cat1, cat2, research_text
    
    # 没有匹配的关键词，尝试根据特定词判断
    if any(w in text_lower for w in ["ic", "芯片", "电路设计", "vlsi"]):
        return "集成电路与芯片", "集成电路设计", research_text
    if any(w in text_lower for w in ["通信", "网络", "传输"]):
        return "通信与网络", "无线通信", research_text
    if any(w in text_lower for w in ["电磁", "微波", "天线"]):
        return "电磁与微波", "射频技术", research_text
    if any(w in text_lower for w in ["人工智能", "机器学习", "深度学习", "图像", "视觉"]):
        return "信号处理与人工智能", "机器学习", research_text
    if any(w in text_lower for w in ["光", "激光", "显示"]):
        return "光电子与显示", "光电子器件", research_text
    if any(w in text_lower for w in ["嵌入式", "系统", "控制"]):
        return "电子系统与嵌入式", "嵌入式系统", research_text
    
    return "其他", "其他", research_text

def get_teacher_list():
    """从列表页面获取所有导师信息"""
    print("正在获取导师列表...")
    try:
        response = requests.get(LIST_URL, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        
        teachers = []
        # 查找表格
        table = soup.find('table')
        if table:
            rows = table.find_all('tr')[1:]  # 跳过表头
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 3:
                    seq = cols[0].get_text(strip=True)
                    name_link = cols[1].find('a')
                    if name_link:
                        name = name_link.get_text(strip=True)
                        url = name_link.get('href', '')
                        if url and not url.startswith('http'):
                            url = 'https://person.zju.edu.cn' + url if url.startswith('/') else 'https://person.zju.edu.cn/' + url
                    else:
                        name = cols[1].get_text(strip=True)
                        url = ''
                    college = cols[2].get_text(strip=True)
                    
                    teachers.append({
                        'seq': seq,
                        'name': name,
                        'college': college,
                        'url': url
                    })
        
        print(f"获取到 {len(teachers)} 位导师")
        return teachers
    except Exception as e:
        print(f"获取导师列表失败: {e}")
        return []

def get_teacher_detail(url, name):
    """获取单个导师的详细信息"""
    if not url:
        return {
            'name': name,
            'title': '未获取',
            'research': '未获取',
            'bio': '未获取',
            'projects': '未获取',
            'email': '未获取'
        }
    
    try:
        print(f"  正在获取: {name} - {url}")
        response = requests.get(url, headers=HEADERS, timeout=30)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取职称
        title = "未获取"
        title_selectors = [
            ('span', {'class': 'title'}),
            ('div', {'class': 'title'}),
            ('p', {'class': 'title'}),
            ('span', {'class': 'job-title'}),
        ]
        for tag, attrs in title_selectors:
            elem = soup.find(tag, attrs)
            if elem:
                title = elem.get_text(strip=True)
                break
        
        # 如果没找到，尝试从页面文本中提取
        if title == "未获取":
            # 尝试从个人信息区域提取
            info_section = soup.find('div', class_='teacher-info') or soup.find('div', class_='person-info')
            if info_section:
                text = info_section.get_text()
                title_match = re.search(r'(教授|副教授|讲师|研究员|副研究员|高级工程师|特聘研究员|青年学者|百人计划|长聘|特聘)', text)
                if title_match:
                    title = title_match.group(1)
            else:
                # 从整个页面提取
                text = soup.get_text()
                title_match = re.search(r'(教授|副教授|讲师|研究员|副研究员|高级工程师|特聘研究员|青年学者|百人计划|长聘|特聘)', text)
                if title_match:
                    title = title_match.group(1)
        
        # 提取研究方向 - 这是重点，需要完整提取
        research = "未获取"
        research_selectors = [
            ('div', {'class': 'research-direction'}),
            ('div', {'class': 'research'}),
            ('div', {'class': 'research_field'}),
            ('div', {'class': 'field'}),
            ('p', {'class': 'research'}),
            ('span', {'class': 'research'}),
        ]
        
        for tag, attrs in research_selectors:
            elems = soup.find_all(tag, attrs)
            if elems:
                research_parts = []
                for elem in elems:
                    text = elem.get_text(strip=True)
                    if text and len(text) > 1:
                        research_parts.append(text)
                if research_parts:
                    research = '；'.join(research_parts)
                    break
        
        # 如果还没找到，尝试通过文本内容查找
        if research == "未获取":
            all_text = soup.find_all(['p', 'div', 'span', 'li'])
            for elem in all_text:
                text = elem.get_text(strip=True)
                # 查找研究方向相关的内容
                if '研究方向' in text or '研究领域' in text or '研究兴趣' in text:
                    # 提取冒号后的内容
                    if '：' in text:
                        parts = text.split('：', 1)
                        if len(parts) > 1 and len(parts[1]) > 2:
                            research = parts[1].strip()
                            break
                    elif ':' in text:
                        parts = text.split(':', 1)
                        if len(parts) > 1 and len(parts[1]) > 2:
                            research = parts[1].strip()
                            break
        
        # 提取个人简介/简历
        bio = "未获取"
        bio_selectors = [
            ('div', {'class': 'bio'}),
            ('div', {'class': 'biography'}),
            ('div', {'class': 'profile'}),
            ('div', {'class': 'introduction'}),
            ('div', {'class': 'teacher-intro'}),
            ('div', {'class': 'person-intro'}),
        ]
        
        for tag, attrs in bio_selectors:
            elem = soup.find(tag, attrs)
            if elem:
                bio = elem.get_text(strip=True)
                break
        
        if bio == "未获取":
            # 尝试查找包含"简介"或"简历"的段落
            all_ps = soup.find_all(['p', 'div'])
            for p in all_ps:
                text = p.get_text(strip=True)
                if ('个人简介' in text or '简历' in text or '教育背景' in text) and len(text) > 20:
                    bio = text
                    break
        
        # 提取项目经历/科研成果
        projects = "未获取"
        proj_selectors = [
            ('div', {'class': 'projects'}),
            ('div', {'class': 'project'}),
            ('div', {'class': 'research-results'}),
            ('div', {'class': 'achievements'}),
            ('div', {'class': 'scientific-research'}),
        ]
        
        for tag, attrs in proj_selectors:
            elem = soup.find(tag, attrs)
            if elem:
                projects = elem.get_text(strip=True)
                break
        
        if projects == "未获取":
            all_divs = soup.find_all('div')
            for div in all_divs:
                text = div.get_text(strip=True)
                if ('科研项目' in text or '科研成果' in text or '主持项目' in text or '承担项目' in text) and len(text) > 20:
                    projects = text
                    break
        
        # 提取邮箱
        email = "未获取"
        # 查找mailto链接
        mailto = soup.find('a', href=re.compile(r'^mailto:'))
        if mailto:
            email = mailto.get('href').replace('mailto:', '')
        else:
            # 从文本中查找邮箱格式
            text = soup.get_text()
            email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', text)
            if email_match:
                email = email_match.group(0)
        
        return {
            'name': name,
            'title': title,
            'research': research,
            'bio': bio,
            'projects': projects,
            'email': email
        }
        
    except Exception as e:
        print(f"    获取 {name} 详情失败: {e}")
        return {
            'name': name,
            'title': '获取失败',
            'research': '获取失败',
            'bio': '获取失败',
            'projects': '获取失败',
            'email': '获取失败'
        }

def save_cv(teacher):
    """保存导师简历到markdown文件"""
    filename = f"{teacher['seq']}_{teacher['name']}.md"
    filepath = os.path.join(CV_DIR, filename)
    
    content = f"""# {teacher['name']}

## 基本信息
- **姓名**: {teacher['name']}
- **学院**: {teacher['college']}
- **职称**: {teacher['title']}
- **邮箱**: {teacher['email']}
- **个人页面**: {teacher['url']}

## 研究方向
{teacher['research']}

## 个人简介
{teacher['bio']}

## 项目经历/科研成果
{teacher['projects']}

---
*数据来源: 浙江大学导师信息系统*
*爬取时间: {time.strftime('%Y-%m-%d %H:%M:%S')}*
"""
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)
    
    return filepath

def create_excel(teachers_data):
    """创建Excel文件，包含多个工作表"""
    print("\n正在创建Excel文件...")
    
    # 准备数据
    all_data = []
    for t in teachers_data:
        cat1, cat2, research = classify_research(t['research'])
        all_data.append({
            '序号': t['seq'],
            '姓名': t['name'],
            '学院': t['college'],
            '职称': t['title'],
            '研究方向大类': cat1,
            '研究方向中类': cat2,
            '具体研究方向': t['research'],
            '项目经历': t['projects'][:200] + '...' if len(t['projects']) > 200 else t['projects'],
            '邮箱': t['email'],
            '个人页面': t['url'],
            '简历文件': f"导师简历/{t['seq']}_{t['name']}.md"
        })
    
    df_all = pd.DataFrame(all_data)
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 定义样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 工作表1: 全部导师
    ws_all = wb.active
    ws_all.title = "全部导师"
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
    
    # 调整列宽
    column_widths = {'A': 6, 'B': 10, 'C': 18, 'D': 12, 'E': 18, 'F': 15, 'G': 50, 'H': 40, 'I': 25, 'J': 35, 'K': 30}
    for col, width in column_widths.items():
        ws_all.column_dimensions[col].width = width
    
    # 工作表: 按大类分表
    categories = df_all['研究方向大类'].unique()
    for cat in sorted(categories):
        df_cat = df_all[df_all['研究方向大类'] == cat]
        ws_cat = wb.create_sheet(title=cat[:20])  # Excel工作表名最多31字符
        for r_idx, row in enumerate(dataframe_to_rows(df_cat, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_cat.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                if r_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
        
        for col, width in column_widths.items():
            ws_cat.column_dimensions[col].width = width
    
    # 工作表: 方向汇总
    ws_summary = wb.create_sheet(title="方向汇总")
    
    # 大类统计
    cat1_counts = df_all['研究方向大类'].value_counts().sort_index()
    ws_summary.cell(row=1, column=1, value="研究方向大类统计").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value="研究方向大类")
    ws_summary.cell(row=2, column=2, value="导师数量")
    ws_summary.cell(row=2, column=3, value="占比")
    
    for i, (cat, count) in enumerate(cat1_counts.items(), 3):
        ws_summary.cell(row=i, column=1, value=cat)
        ws_summary.cell(row=i, column=2, value=count)
        ws_summary.cell(row=i, column=3, value=f"{count/len(df_all)*100:.1f}%")
    
    # 中类统计
    start_row = len(cat1_counts) + 5
    ws_summary.cell(row=start_row, column=1, value="研究方向中类统计").font = Font(bold=True, size=14)
    ws_summary.cell(row=start_row+1, column=1, value="研究方向大类")
    ws_summary.cell(row=start_row+1, column=2, value="研究方向中类")
    ws_summary.cell(row=start_row+1, column=3, value="导师数量")
    
    cat2_counts = df_all.groupby(['研究方向大类', '研究方向中类']).size().reset_index(name='count')
    for i, row in enumerate(cat2_counts.itertuples(), start_row+2):
        ws_summary.cell(row=i, column=1, value=row[1])  # 大类
        ws_summary.cell(row=i, column=2, value=row[2])  # 中类
        ws_summary.cell(row=i, column=3, value=row[3])  # 数量
    
    for col in ['A', 'B', 'C']:
        ws_summary.column_dimensions[col].width = 25
    
    # 保存
    wb.save(EXCEL_PATH)
    print(f"Excel文件已保存: {EXCEL_PATH}")

def generate_report(teachers_data):
    """生成分析报告"""
    print("\n正在生成分析报告...")
    
    # 统计数据
    total = len(teachers_data)
    df = pd.DataFrame([{
        'name': t['name'],
        'college': t['college'],
        'title': t['title'],
        'research': t['research'],
        'cat1': classify_research(t['research'])[0],
        'cat2': classify_research(t['research'])[1]
    } for t in teachers_data])
    
    # 学院分布
    college_counts = df['college'].value_counts()
    
    # 职称分布
    title_counts = df['title'].value_counts()
    
    # 方向分布
    cat1_counts = df['cat1'].value_counts()
    cat2_counts = df.groupby(['cat1', 'cat2']).size().reset_index(name='count')
    
    report = f"""# 浙江大学电子信息技术方向导师分析报告

> 生成时间: {time.strftime('%Y年%m月%d日')}
> 数据来源: https://pi.zju.edu.cn/2026/0317/c90502a3141478/page.htm
> 导师总数: {total}位

---

## 一、学院分布分析

| 学院 | 导师数量 | 占比 |
|------|---------|------|
"""
    for college, count in college_counts.items():
        report += f"| {college} | {count} | {count/total*100:.1f}% |\n"
    
    report += f"""
### 学院分布特点

- **信息与电子工程学院**是主力学院，拥有{college_counts.get('信息与电子工程学院', 0)}位导师，占比{college_counts.get('信息与电子工程学院', 0)/total*100:.1f}%
- **集成电路学院**作为新兴学院，拥有{college_counts.get('集成电路学院', 0)}位导师，体现学校对集成电路领域的重视

---

## 二、职称结构分析

| 职称 | 数量 | 占比 |
|------|------|------|
"""
    for title, count in title_counts.items():
        report += f"| {title} | {count} | {count/total*100:.1f}% |\n"
    
    report += f"""
### 职称结构特点

- **高级职称占比高**：教授、研究员等正高级职称占比较高，师资力量雄厚
- **青年学者活跃**：百人计划、特聘研究员等青年学者比例可观，体现队伍活力

---

## 三、研究方向深度分析（重点）

### 3.1 大类分布统计

| 研究方向大类 | 导师数量 | 占比 |
|-------------|---------|------|
"""
    for cat, count in cat1_counts.sort_values(ascending=False).items():
        report += f"| {cat} | {count} | {count/total*100:.1f}% |\n"
    
    report += """
### 3.2 各子领域分布详情

"""
    for cat1 in sorted(df['cat1'].unique()):
        cat2_in_cat1 = cat2_counts[cat2_counts['cat1'] == cat1].sort_values('count', ascending=False)
        report += f"#### {cat1}\n\n"
        report += "| 子领域 | 导师数量 |\n|--------|----------|\n"
        for _, row in cat2_in_cat1.iterrows():
            report += f"| {row['cat2']} | {row['count']} |\n"
        report += "\n"
    
    # 护城河分析
    report += """
### 3.3 "护城河"方向分析

基于技术壁垒、人才稀缺度、产业需求、薪资水平和发展前景，以下是具有"护城河"优势的研究方向：

#### 🏆 高护城河方向（推荐）

| 方向 | 护城河指数 | 理由 |
|------|-----------|------|
| **集成电路设计（模拟IC/射频IC）** | ⭐⭐⭐⭐⭐ | 技术壁垒极高，人才极度稀缺，薪资顶尖，国家战略需求 |
| **半导体器件（氮化镓/碳化硅）** | ⭐⭐⭐⭐⭐ | 第三代半导体国家战略，产业需求爆发，技术门槛高 |
| **EDA工具开发** | ⭐⭐⭐⭐⭐ | 被卡脖子领域，国产化需求迫切，人才极度稀缺 |
| **射频技术与天线（毫米波/相控阵）** | ⭐⭐⭐⭐ | 5G/6G核心，技术难度大，应用领域广 |
| **AI芯片设计** | ⭐⭐⭐⭐ | 人工智能浪潮核心，创业机会多，薪资高 |

#### 🔥 热门且有前景方向

| 方向 | 热度指数 | 理由 |
|------|---------|------|
| **深度学习/机器学习** | ⭐⭐⭐⭐⭐ | 产业需求大，应用广泛，但竞争也激烈 |
| **光电子器件** | ⭐⭐⭐⭐ | 光通信、激光器需求旺盛，技术积累价值高 |
| **MEMS传感器** | ⭐⭐⭐⭐ | 物联网时代基础，应用广泛，技术壁垒适中 |

#### 💰 就业前景与经济回报分析

**第一梯队（年薪50万+）**：
- 模拟IC设计、射频IC设计、AI芯片架构
- EDA工具开发、先进封装
- 顶尖企业：华为海思、紫光展锐、阿里平头哥、字节AI Lab

**第二梯队（年薪30-50万）**：
- 数字IC设计、半导体器件研发
- 射频工程师、天线工程师
- 就业去向：芯片公司、通信设备商、互联网公司

**第三梯队（年薪20-30万）**：
- 嵌入式系统、信号处理
- 图像处理、通信算法
- 就业面广，选择多

---

## 四、选导师策略建议

### 4.1 职业导向选择

**目标：高薪芯片行业**
- 优先选择：集成电路设计、半导体器件方向
- 推荐导师特征：有产业合作项目、与企业联合培养
- 关键技能：EDA工具熟练、有流片经验

**目标：学术科研道路**
- 优先选择：有国家自然科学基金、发表顶会论文的导师
- 推荐方向：前沿交叉方向、国家自然科学基金重点支持领域
- 关键技能：论文写作、项目申请

**目标：互联网/AI行业**
- 优先选择：机器学习、计算机视觉方向
- 推荐导师特征：与企业有合作、提供实习机会
- 关键技能：编程能力、算法能力

### 4.2 导师类型选择

**资深教授**
- 优点：资源丰富、人脉广、项目多
- 注意：可能较忙、指导时间有限

**青年学者（百人/特聘）**
- 优点：精力充沛、指导细致、上升期
- 注意：资源积累中、压力大

**产业背景导师**
- 优点：产业资源、实习机会、就业推荐
- 注意：学术发表可能相对少

### 4.3 实用建议

1. **提前联系**：热门导师名额有限，尽早联系
2. **了解风格**：通过师兄师姐了解导师风格
3. **查看项目**：关注导师近3年的项目和论文
4. **实地考察**：有机会到实验室参观交流

---

## 五、导师简历链接

### 按研究方向大类

"""
    for cat1 in sorted(df['cat1'].unique()):
        report += f"#### {cat1}\n\n"
        cat1_teachers = [t for t in teachers_data if classify_research(t['research'])[0] == cat1]
        for t in cat1_teachers[:10]:  # 每类显示前10个
            cv_link = f"导师简历/{t['seq']}_{t['name']}.md"
            report += f"- [{t['name']}]({cv_link}) - {t['college']} - {classify_research(t['research'])[1]}\n"
        if len(cat1_teachers) > 10:
            report += f"- *...还有{len(cat1_teachers)-10}位导师*\n"
        report += "\n"
    
    report += f"""
---

## 附录：数据说明

- 数据爬取时间：{time.strftime('%Y年%m月%d日 %H:%M:%S')}
- 数据来源：浙江大学导师信息系统
- 分析方法：基于研究方向关键词进行智能分类
- 注意：部分导师信息可能因网页结构差异而未完整获取，建议以导师个人主页信息为准

---

*本报告由AI自动生成，仅供参考*
"""
    
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"分析报告已保存: {REPORT_PATH}")

def main():
    """主函数"""
    print("=" * 60)
    print("浙江大学电子信息技术方向导师信息爬取")
    print("=" * 60)
    
    # 1. 获取导师列表
    teachers = get_teacher_list()
    if not teachers:
        print("获取导师列表失败，退出")
        return
    
    # 2. 获取每位导师的详细信息
    print(f"\n开始爬取 {len(teachers)} 位导师的详细信息...")
    teachers_data = []
    
    for i, t in enumerate(teachers):
        print(f"[{i+1}/{len(teachers)}] 处理: {t['name']}")
        
        # 获取详细信息
        detail = get_teacher_detail(t['url'], t['name'])
        
        teacher_data = {
            'seq': t['seq'],
            'name': t['name'],
            'college': t['college'],
            'url': t['url'],
            **detail
        }
        teachers_data.append(teacher_data)
        
        # 保存简历
        save_cv(teacher_data)
        
        # 控制请求频率
        time.sleep(1.5)
    
    # 3. 创建Excel文件
    create_excel(teachers_data)
    
    # 4. 生成分析报告
    generate_report(teachers_data)
    
    print("\n" + "=" * 60)
    print("爬取完成！")
    print(f"Excel文件: {EXCEL_PATH}")
    print(f"简历目录: {CV_DIR}")
    print(f"分析报告: {REPORT_PATH}")
    print("=" * 60)

if __name__ == "__main__":
    main()